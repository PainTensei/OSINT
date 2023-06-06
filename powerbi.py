# for power bi:
# py -m pip install pandas, matplotlib, requests, shodan, python-whois, beautifulsoup4
import socket
import urllib.parse
import subprocess
try:import requests
except ImportError:
    subprocess.call(['pip', 'install', 'requests'])
import requests
try:import shodan
except ImportError:
    subprocess.call(['pip', 'install', 'shodan'])
import shodan
try:import whois
except ImportError:
    subprocess.call(['pip', 'install', 'python-whois'])
import whois
try:import bs4
except ImportError:
    subprocess.call(['pip', 'install', 'beautifulsoup4'])
from bs4 import BeautifulSoup
try:import easygui
except ImportError:
    subprocess.call(['pip', 'install', 'easygui'])
import easygui
import openpyxl

titlescreen = """
\033[1;31;40m  ,---.   ,-----.  ,---.          ,-----.  ,---.  ,--.,--.  ,--.,--------.
\033[1;31;40m /  O  \ '  .--./ /  O  \ ,-----.'  .-.  ''   .-' |  ||  ,'.|  |'--.  .--'
\033[1;37;40m|  .-.  ||  |    |  .-.  |'-----'|  | |  |`.  `-. |  ||  |' '  |   |  |
\033[1;34;40m|  | |  |'  '--'\|  | |  |       '  '-'  '.-'    ||  ||  | `   |   |  |
\033[1;34;40m`--' `--' `-----'`--' `--'        `-----' `-----' `--'`--'  `--'   `--'\033[0m
"""
print(titlescreen)

# target popup
target = easygui.enterbox("Enter the website URL: ")

# Target
url = target
parsed_url = urllib.parse.urlparse("https://" + url)
domain_name = parsed_url.netloc

# IP
ip_address = socket.gethostbyname(domain_name)
print(f"\nIP Address: {ip_address}")

# WHOIS
domain_info = whois.whois(domain_name)

# DNS
print(f"DNS Information:")
for dns in domain_info.name_servers:
    print(f"\t{dns}")

# Crawler
page = requests.get("https://" + url)
soup = BeautifulSoup(page.content, 'html.parser')
emails = set()
usernames = set()
for link in soup.find_all('a'):
    href = link.get('href')
    if href is not None and ('mailto:' in href):
        email = href.split(':')[-1]
        emails.add(email)
    elif href is not None and ('twitter.com/' in href or 'instagram.com/' in href):
        username = href.split('/')[-1]
        usernames.add(username)

# Emails & usernames
if emails:
    print("Emails:")
    for email in emails:
        print(f"\t{email}")
else: print("No emails found on the webpage.")

if usernames:
    print("Usernames:")
    for username in usernames:
        print(f"\t{username}")
else: print("No usernames found on the webpage.")

print(f"all hyperlinks found:\t")
for link in soup.find_all('a'):
    href = link.get('href')
    if href is not None and ('http' in href):print(f"\t{href}")

# Wayback Machine
wayback_json = f"http://archive.org/wayback/available?url={url}"
response = requests.get(wayback_json)
data = response.json()
    # Check if snapshots are saved
if 'closest' not in data['archived_snapshots']:
    print(f"The target '{url}' is not in the Wayback Machine")
else:
    # Scrape the number of snapshots found
    page2 = requests.get("http://web.archive.org/cdx/search/cdx?url=" + url + "&output=json")
    soup2 = BeautifulSoup(page2.content, 'html.parser')
    number_wayback = page2.text.count(url)
    print(f"\nNumber of snapshots found: {number_wayback}")
    # Get the URL of the most recent archived snapshot
    snapshot_url = data['archived_snapshots']['closest']['url']
    print(f"Most recent snapshot: {snapshot_url}")

# Shodan
# api_key = "LEehS1398exXggIEL2NMw2RduwkLZbJZ"
# api = shodan.Shodan(api_key)
# results = api.host(ip_address)
# print(results)

# Dorks

# ------------------ creating a report file ------------------

# Create a new workbook
workbook = openpyxl.Workbook()

# Get the active worksheet
worksheet = workbook.active

# Add headers to the worksheet
worksheet['A1'] = 'IP Address'
worksheet['B1'] = 'DNS'
worksheet['C1'] = 'Emails'
worksheet['D1'] = 'Usernames'
worksheet['E1'] = 'Hyperlinks'
worksheet['F1'] = 'Snapshots'
worksheet['G1'] = 'Most recent snapshot URL:'

# Write the data to the worksheet
worksheet.cell(row=2, column=1, value=ip_address)
row = 2
for dns in domain_info.name_servers:
    worksheet.cell(row=row, column=2, value=dns)
    row += 1
row2 = 2
for email in emails:
    worksheet.cell(row=row2, column=3, value=email)
    row2 += 1
row3 = 2
for username in usernames:
    worksheet.cell(row=row3, column=4, value=username)
    row3 += 1
row4 = 2
for link in soup.find_all('a'):
    href = link.get('href')
    if href is not None and ('http' in href):
        worksheet.cell(row=row4, column=5, value=href)
        row4 += 1
worksheet.cell(row=2, column=6, value=number_wayback)
worksheet.cell(row=2, column=7, value=snapshot_url)

# Save the workbook
workbook.save(f"{domain_name}.xlsx")
